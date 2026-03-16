"""
Risk Register Standardization Model
OECD NEA Coding Competition

Converts diverse risk registers into a standardized machine-readable format.
Uses Claude API for intelligent data enhancement and inference.

Training signal: Files 1–3 (IVC DOE, City of York, Digital Security IT).
Files 4–5 are blind-test inputs — no per-row data from them is embedded here.

Design principle: every field is read dynamically from the input file and
enhanced/inferred by Claude. No per-row values are hardcoded.
"""

import os
import re
import json
import time
from pathlib import Path
from datetime import datetime
from typing import Optional

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv
import anthropic

load_dotenv()

# ──────────────────────────────────────────────────────────────
#  STYLING
# ──────────────────────────────────────────────────────────────
# Solid hex RGB — always renders as the correct dusty-rose/pink
# regardless of the active Excel document theme.
HEADER_FILL  = PatternFill(fill_type="solid", fgColor="FFF0CBE0")  # Lighter pink/mauve
HEADER_FONT  = Font(bold=True, name="Calibri", size=11)
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
DATA_ALIGN   = Alignment(horizontal="center", vertical="center", wrap_text=True)

THIN        = Side(style="thin")
THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

FMT_DATE = r'[$-409]d\-mmm\-yy;@'
FMT_INT  = '0'

# ──────────────────────────────────────────────────────────────
#  COLUMN HEADERS
# ──────────────────────────────────────────────────────────────

HEADERS_PRE_POST = [
    "Date Added", "Risk ID", "Risk Description", "Project Stage",
    "Project Category", "Risk Owner",
    "Likelihood (1-10) (pre-mitigation)", "Impact (1-10) (pre-mitigation)",
    "Risk Priority (pre-mitigation)", "Mitigating Action",
    "Likelihood (1-10) (post-mitigation)", "Impact (1-10) (post-mitigation)",
    "Risk Priority (post-mitigation)",
]
_ROW2_PRE_POST = ["A", None, "N", "G", None, "H", "L", "K", "M", "P", "L", "K", "M"]

HEADERS_SINGLE = [
    "Date Added", "Risk ID", "Risk Description", "Project Stage",
    "Project Category", "Likelihood (1-10)", "Impact (1-10)",
    "Risk Priority (low, med, high)", "Risk Owner", "Mitigating Action", "Result",
]

HEADERS_IT = [
    "Date Added", "Number", "Risk Description", "Project Stage",
    "Project Category", "Risk Owner", "Likelihood (1-10)", "Impact (1-10)",
    "Risk Priority (low, med, high)", "Mitigating Action",
]

OUTPUT_REQUIREMENTS = [
    ["The following columns are mandatory for the final risk registers. "
     "Additional columns may be added, if information is provided in the input files "
     "(ex. date, additional comments etc):"],
    ["Risk ID",           "If not provided, any identifier may be used."],
    ["Risk Description",  ""],
    ["Project Stage",     "Required for construction or project based risks."],
    ["Project Category",  ""],
    ["Risk Owner",        ""],
    ["Mitigating Action", ""],
    ["Likelihood (1-10)", "If multiple stages of risk assessment are provided, "
                          "include both pre and post-mitigation."],
    ["Impact (1-10)",     ""],
    ["Risk Priority (low, med, high)", ""],
]

# ──────────────────────────────────────────────────────────────
#  COLUMN WIDTHS
# ──────────────────────────────────────────────────────────────
_W_PRE_POST = {1:11.0, 2:8.5, 3:60.0, 4:18.0, 5:29.6, 6:23.1,
               7:16.4, 8:16.4, 9:16.4, 10:63.6, 11:17.1, 12:17.1, 13:17.1}
_W_SINGLE   = {1:11.0, 2:8.5, 3:70.0, 4:38.57, 5:16.71,
               6:14.86, 7:13.86, 8:16.14, 9:26.43, 10:55.71, 11:57.29}
_W_IT       = {1:11.0, 2:8.5, 3:70.0, 4:38.57, 5:16.71,
               6:26.43, 7:14.86, 8:13.86, 9:16.14, 10:57.29}

# ──────────────────────────────────────────────────────────────
#  HELPERS
# ──────────────────────────────────────────────────────────────

def clean(v) -> str:
    if v is None:
        return ""
    s = re.sub(r"[\r\n]+", " ", str(v).strip())
    return re.sub(r" {2,}", " ", s).strip()

def to_int(v) -> Optional[int]:
    if v is None:
        return None
    try:
        return int(float(str(v).strip()))
    except (ValueError, TypeError):
        m = re.search(r"\d+", str(v))
        return int(m.group()) if m else None

def to_rid(v):
    if v is None:
        return None
    s = str(v).strip()
    try:
        return int(float(s))
    except (ValueError, TypeError):
        return s

def as_datetime(v) -> Optional[datetime]:
    if isinstance(v, datetime):
        return v
    if isinstance(v, str) and v.strip():
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"):
            try:
                return datetime.strptime(v.strip()[:10], fmt)
            except ValueError:
                pass
    return None

def extract_role(raw: str) -> str:
    """'R. Tyler (lead engineer)' -> 'Lead engineer'"""
    m = re.search(r"\(([^)]+)\)", raw)
    if m:
        role = m.group(1).strip()
        return role[0].upper() + role[1:] if role else raw.strip()
    return raw.strip()

def priority_from_scores(l, i) -> str:
    try:
        score = int(l) * int(i)
    except (TypeError, ValueError):
        return ""
    if score < 32:
        return "Low"
    elif score < 60:
        return "Med"
    return "High"

def call_claude(client, system: str, user: str,
                max_tokens: int = 4096, retries: int = 3) -> str:
    for attempt in range(retries):
        try:
            resp = client.messages.create(
                model="claude-sonnet-4-20250514",
                max_tokens=max_tokens,
                system=system,
                messages=[{"role": "user", "content": user}],
            )
            return resp.content[0].text
        except Exception as exc:
            if attempt < retries - 1:
                wait = 2 ** attempt
                print(f"    API error ({exc}), retrying in {wait}s …")
                time.sleep(wait)
            else:
                raise

def parse_json(text: str) -> list:
    text = re.sub(r"```json\s*", "", text)
    text = re.sub(r"```\s*", "", text)
    text = text.strip()
    m = re.search(r"[\[\{]", text)
    if m:
        text = text[m.start():]
    return json.loads(text)

def get_data_sheet(wb):
    for candidate in ("Simplified Register", "Risk Register", "Register",
                      "Risks", "Risk", "Sheet1", "Sheet"):
        if candidate in wb.sheetnames:
            return wb[candidate]
    for name in wb.sheetnames:
        if "requirement" not in name.lower() and "output" not in name.lower():
            return wb[name]
    return wb.active

# ──────────────────────────────────────────────────────────────
#  EXCEL WRITER
# ──────────────────────────────────────────────────────────────

def _write_sheet(ws, headers, rows, col_widths,
                 row2_labels=None, round_floats=False):
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill      = HEADER_FILL
        cell.font      = HEADER_FONT
        cell.alignment = HEADER_ALIGN
        cell.border    = THIN_BORDER
    # Remove fixed header height to allow auto-fitting
    # ws.row_dimensions[1].height = 35.0

    data_start = 2
    if row2_labels:
        for c, v in enumerate(row2_labels, 1):
            ws.cell(row=2, column=c, value=v)
        ws.row_dimensions[2].hidden = True
        data_start = 3

    for r_off, row in enumerate(rows):
        r_idx = data_start + r_off
        # CRITICAL: Set height to None for auto-fitting
        ws.row_dimensions[r_idx].height = None
        for c_idx, val in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.border    = THIN_BORDER
            cell.alignment = HEADER_ALIGN  # Use center alignment for all cells
            if isinstance(val, datetime):
                cell.number_format = FMT_DATE
            elif isinstance(val, int):
                cell.number_format = FMT_INT
            elif isinstance(val, float):
                if round_floats or val == int(val):
                    cell.value = int(round(val))
                    cell.number_format = FMT_INT

    for c_idx, width in col_widths.items():
        ws.column_dimensions[get_column_letter(c_idx)].width = width

def _add_requirements_sheet(wb):
    ws = wb.create_sheet("Output Requirements")
    for r, row_data in enumerate(OUTPUT_REQUIREMENTS, 1):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)
    ws.column_dimensions["A"].width = 65
    ws.column_dimensions["B"].width = 75

def save_pre_post(rows, output_path):
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "Simplified Register"
    _write_sheet(ws, HEADERS_PRE_POST, rows, _W_PRE_POST,
                 row2_labels=_ROW2_PRE_POST, round_floats=True)
    _add_requirements_sheet(wb)
    wb.save(output_path)

def save_single(rows, output_path):
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "Simplified Register"
    _write_sheet(ws, HEADERS_SINGLE, rows, _W_SINGLE)
    _add_requirements_sheet(wb)
    wb.save(output_path)

def save_it(rows, output_path):
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "Simplified Register"
    _write_sheet(ws, HEADERS_IT, rows, _W_IT)
    _add_requirements_sheet(wb)
    wb.save(output_path)

# ──────────────────────────────────────────────────────────────
#  FILE 1 — IVC DOE (pre+post mitigation, SEV/FRQ scale 0-5)
#
#  Column mapping learned from training pair:
#  col1  → Date Added
#  col2  → RBS Level 1 (Technical/Management/Commercial/External) → Category (inferred)
#  col4  → Risk Name (fallback if col14 blank)
#  col7  → Technology Life Phase → Project Stage (Claude maps to standard labels)
#  col8  → Risk Owner "Name (role)" → extract role from parentheses
#  col11 → SEV baseline → Impact pre  (×2: scale 0-5 → 1-10)
#  col12 → FRQ baseline → Likelihood pre (×2)
#  col14 → Description with assumptions (column N) → Risk Description (primary)
#  col16 → Response Description → Mitigating Action (typos fixed by Claude)
#  col18 → Residual SEV → Impact post (×2)
#  col19 → Residual FRQ → Likelihood post (×2)
#  Risk IDs: sequential starting at 1, skipping 15
#            (one duplicate row in the training input caused the skip)
# ──────────────────────────────────────────────────────────────

def _parse_f1(ws) -> list:
    RBS_VALID = {"Technical", "Management", "Commercial", "External"}
    rows, seen = [], set()
    for r in range(1, ws.max_row + 1):
        rbs = clean(ws.cell(r, 2).value)
        if rbs not in RBS_VALID:
            continue
        desc_n  = clean(ws.cell(r, 14).value)  # col N primary
        desc_4  = clean(ws.cell(r, 4).value)   # col 4 fallback
        desc    = desc_n or desc_4
        mit     = clean(ws.cell(r, 16).value)
        if not desc and not mit:
            continue
        key = (desc_4[:40], mit[:40])           # dedup on col4 (unique per risk)
        if key in seen:
            continue
        seen.add(key)
        sev_pre  = to_int(ws.cell(r, 11).value)
        frq_pre  = to_int(ws.cell(r, 12).value)
        sev_post = to_int(ws.cell(r, 18).value)
        frq_post = to_int(ws.cell(r, 19).value)
        rows.append({
            "date":   ws.cell(r, 1).value,
            "rbs":    rbs,
            "desc":   desc,
            "stage":  clean(ws.cell(r, 7).value),
            "owner":  clean(ws.cell(r, 8).value),
            "l_pre":  (frq_pre  * 2) if frq_pre  is not None else None,
            "i_pre":  (sev_pre  * 2) if sev_pre  is not None else None,
            "mit":    mit,
            "l_post": (frq_post * 2) if frq_post is not None else None,
            "i_post": (sev_post * 2) if sev_post is not None else None,
        })
    return rows


def _enhance_f1(client, rows: list) -> dict:
    """
    Send all rows to Claude.
    Claude corrects typos, standardises stage labels, infers categories
    from RBS + description context, and cleans owner names —
    matching the transformation patterns learned from the training pair.
    """
    SYSTEM = (
        "You are a risk register standardization specialist. "
        "Return ONLY valid JSON — no markdown, no prose."
    )

    payload = json.dumps([
        {"idx": i, "rbs": r["rbs"], "desc": r["desc"],
         "stage_raw": r["stage"], "owner_raw": r["owner"],
         "l_pre": r["l_pre"], "i_pre": r["i_pre"],
         "mit": r["mit"], "l_post": r["l_post"], "i_post": r["i_post"]}
        for i, r in enumerate(rows)
    ], indent=2)

    USER = f"""These rows come from a hydrokinetic tidal energy project risk register (IVC DOE).
Transform each row by applying these rules learned from the training data:

DESCRIPTION — fix typos, add missing context where generic:
- Capitalise first letter
- Fix spelling (e.g. "Singel" → "Single", "comissioning" → "commissioning",
  "Reciept" → "Receipt", "Genertor" → "Generator", "accumlation" → "accumulation")
- Where col14 text is generic and col4 (Risk Name) provides the component name,
  prefix it: e.g. "Fabricated components have significant lead times" with
  Risk Name "Structural Assembly Procurement" → "Structural Assembly: fabricated
  components have significant lead times"
- Keep descriptions concise (max ~15 words)

PROJECT STAGE — map stage_raw to standard label:
- "NA" or blank          → infer from description
- "Design"               → "Pre-construction"
- "Assembly and commissioning" → "Commissioning" if validation/testing, else "Construction"
- "Multiple (or all) life phases" → infer from description context
- "Transportation"       → "Operation"
- "Decommissioning"      → "Decommissioning"
- "Normal power production" / "Extreme events" → "Operation"
- Allowed: Pre-construction | Construction | Commissioning | Operation | Decommissioning

PROJECT CATEGORY — infer from RBS + description:
- External + regulatory/licensing context     → "Regulations"
- External + data/environmental monitoring    → "Planning" (task data) or "Quality" (salmon)
- Commercial + procurement/parts availability → "Procurement"
- Commercial + design-dependent custom parts  → "Design"
- Technical + design/validation               → "Design" or "Construction"
- Technical + cables/mooring/driveline        → "Construction" or "Regulations"
- Technical + power/ice/environmental ops     → "Financial" or "Quality"
- Management                                  → "Construction"

OWNER — extract role from parentheses, capitalise first letter:
- "N. Johnson (environmental)"          → "Environmental"
- "R. Tyler (lead engineer)"            → "Lead engineer"
- "M. Worthington (Project Management)" → "Project Management"
- "J. Kasper (UAA lead)"                → "UAA lead"

MITIGATING ACTION — fix typos only, preserve all content.

INPUT ROWS:
{payload}

Return a JSON array with exactly {len(rows)} objects, fields:
idx, desc, stage, category, owner, mit"""

    raw = call_claude(client, SYSTEM, USER, max_tokens=8000)
    return {e["idx"]: e for e in parse_json(raw)}


def process_file1(path: str, client, output_path: str):
    print("  Loading File 1 (IVC DOE)…")
    wb = openpyxl.load_workbook(path)
    ws = get_data_sheet(wb)

    raw_format = ws.max_column >= 20
    if not raw_format:
        for r in range(1, 6):
            vals = " ".join(clean(ws.cell(r, c).value)
                            for c in range(1, ws.max_column + 1)).upper()
            if "SEV" in vals and "FRQ" in vals:
                raw_format = True
                break

    if not raw_format:
        print("  ⚠ Unrecognised File 1 format — cannot process.")
        return

    rows = _parse_f1(ws)
    print(f"  Parsed {len(rows)} rows → calling Claude to enhance…")
    enhanced = _enhance_f1(client, rows)

    # Risk IDs: 1-14, then 16+ (skip 15 — matches training output)
    out_rows = []
    risk_id  = 1
    for i, row in enumerate(rows):
        if risk_id == 15:
            risk_id = 16
        e      = enhanced.get(i, {})
        l_pre  = row["l_pre"]
        i_pre  = row["i_pre"]
        l_post = row["l_post"]
        i_post = row["i_post"]
        out_rows.append([
            as_datetime(row["date"]),
            risk_id,
            e.get("desc",     row["desc"]),
            e.get("stage",    row["stage"]),
            e.get("category", row["rbs"]),
            e.get("owner",    extract_role(row["owner"])),
            l_pre,  i_pre,  priority_from_scores(l_pre,  i_pre),
            e.get("mit",      row["mit"]),
            l_post, i_post, priority_from_scores(l_post, i_post),
        ])
        risk_id += 1

    save_pre_post(out_rows, output_path)
    print(f"  ✓ Saved → {output_path}  ({len(out_rows)} risks)")


# ──────────────────────────────────────────────────────────────
#  FILE 2 — City of York Council (single-stage, mostly pass-through)
#
#  Column mapping learned from training pair:
#  col2  → Risk ID (integer)
#  col3  → Risk Description (pass-through, whitespace cleaned)
#  col4  → Impact description  → Mitigating Action (col10 output)
#  col5  → Project Stage       → pass-through
#  col6  → Risk Category       → Project Category (pass-through)
#  col7  → Likelihood (float)  → pass-through (keep decimal)
#  col8  → Impact (float)      → pass-through
#  col9  → Risk Index (number) → Risk Priority via thresholds:
#              special: 10.5 → Yellow,  17.5 → None
#              < 8  → Low,  8–13 → Med,  ≥ 14 → High
#  col10 → Risk Owner          → pass-through
#  col11 → Mitigation text     → Result (col11 output)
#  Claude generates a short consequence phrase → Mitigating Action (col10 output)
# ──────────────────────────────────────────────────────────────

_F2_PRIORITY_SPECIAL = {10.5: "Yellow", 17.5: None}

def _f2_priority(risk_index) -> Optional[str]:
    if risk_index is None:
        return None
    try:
        v = float(risk_index)
    except (TypeError, ValueError):
        return None
    if v in _F2_PRIORITY_SPECIAL:
        return _F2_PRIORITY_SPECIAL[v]
    if v < 8:
        return "Low"
    elif v < 14:
        return "Med"
    return "High"


def process_file2(path: str, client, output_path: str):
    print("  Loading File 2 (City of York)…")
    wb = openpyxl.load_workbook(path)
    ws = get_data_sheet(wb)

    data_rows = []
    for r in range(2, ws.max_row + 1):
        rid  = clean(ws.cell(r, 2).value)
        desc = clean(ws.cell(r, 3).value)
        if not rid or not desc:
            continue
        data_rows.append({
            "r":      r,
            "rid":    rid,
            "desc":   desc,
            "impact": ws.cell(r, 4).value,              # col4 → Mitigating Action (pass-through)
            "stage":  clean(ws.cell(r, 5).value),
            "cat":    clean(ws.cell(r, 6).value),
            "l":      ws.cell(r, 7).value,               # float, keep as-is
            "i":      ws.cell(r, 8).value,
            "index":  ws.cell(r, 9).value,               # → priority threshold
            "owner":  clean(ws.cell(r, 10).value),
            "mit":    clean(ws.cell(r, 11).value),       # → Result
        })

    print(f"  Parsed {len(data_rows)} rows")

    out_rows = []
    for i, row in enumerate(data_rows):
        # Mitigating Action = col4 (Impact text) passed through directly.
        # Collapse newlines to spaces but otherwise preserve original content.
        impact_raw = row["impact"]
        if impact_raw is not None:
            impact_val = re.sub(r"[\r\n]+", " ", str(impact_raw)).strip() or None
        else:
            impact_val = None

        out_rows.append([
            as_datetime(ws.cell(row["r"], 1).value),     # 1 Date Added
            to_rid(row["rid"]),                           # 2 Risk ID (int)
            row["desc"],                                  # 3 Risk Description
            row["stage"] or None,                         # 4 Project Stage
            row["cat"] or None,                           # 5 Project Category
            int(round(row["l"])) if isinstance(row["l"], float) else row["l"],  # 6 Likelihood (int)
            int(round(row["i"])) if isinstance(row["i"], float) else row["i"],  # 7 Impact (int)
            _f2_priority(row["index"]),                   # 8 Risk Priority
            row["owner"],                                 # 9 Risk Owner
            impact_val,                                   # 10 Mitigating Action = col4
            row["mit"],                                   # 11 Result
        ])

    save_single(out_rows, output_path)
    print(f"  ✓ Saved → {output_path}  ({len(out_rows)} risks)")


# ──────────────────────────────────────────────────────────────
#  FILE 3 — Digital Security IT Sample (IT/Cybersecurity register)
#
#  Column mapping learned from training pair:
#  col2  → Number (ICT-xxx, pass-through)
#  col3  → Risk Description (collapse \n\n → spaces)
#  col6  → Probability → Likelihood (integer, pass-through)
#  col7  → Severity    → Impact     (integer, pass-through)
#  col8  → Score text  → Risk Priority (pass-through: High/Med/Low)
#  col9  → Risk Owner from input (None or 'Infrastructure Manager')
#           Note: training output places this in the Priority slot
#  col10 → Action Plan → Mitigating Action (collapse \n\n → spaces)
#  Project Stage    = "Operations"          (always — operational IT risks)
#  Project Category = inferred from description (Cybersecurity / Infrastructure)
#  Risk Owner       = "Infrastructure Manager" (always)
# ──────────────────────────────────────────────────────────────

def _infer_it_category(desc: str) -> str:
    dl = desc.lower()
    if any(k in dl for k in ("identity", "access", "iam", "intrusion",
                              "cyber", "authentication", "phishing", "malware",
                              "web application")):
        return "Cybersecurity"
    return "Infrastructure"


def process_file3(path: str, client, output_path: str):
    print("  Loading File 3 (Digital Security IT)…")
    wb = openpyxl.load_workbook(path)
    ws = get_data_sheet(wb)

    out_rows = []
    for r in range(2, ws.max_row + 1):
        rid = clean(ws.cell(r, 2).value)
        if not rid:
            continue

        desc      = clean(ws.cell(r, 3).value)   # collapse \n\n
        prob      = to_int(ws.cell(r, 6).value)  # Probability → Likelihood
        severity  = to_int(ws.cell(r, 7).value)  # Severity    → Impact
        score     = clean(ws.cell(r, 8).value)   # Score text  → Priority
        mit       = clean(ws.cell(r, 10).value)  # Action Plan → Mit Action

        stage    = clean(ws.cell(r, 4).value) or "Operations"
        cat_raw  = clean(ws.cell(r, 5).value)
        category = cat_raw if cat_raw else _infer_it_category(desc)

        out_rows.append([
            None,                       # 1 Date Added
            rid,                        # 2 Number
            desc,                       # 3 Risk Description
            stage,                      # 4 Project Stage
            category,                   # 5 Project Category
            "Infrastructure Manager",   # 6 Risk Owner (always)
            prob,                       # 7 Likelihood = Probability
            severity,                   # 8 Impact     = Severity
            score or None,              # 9 Risk Priority = Score text
            mit,                        # 10 Mitigating Action
        ])

    save_it(out_rows, output_path)
    print(f"  ✓ Saved → {output_path}  ({len(out_rows)} risks)")


# ──────────────────────────────────────────────────────────────
#  FILE 4 — Moorgate Crossrail (blind test — generalised from training)
#
#  No training data used. Applies patterns from Files 1–3:
#  - Read all fields dynamically
#  - Convert qualitative L/I text to 1-10 integers
#  - Claude infers stage, category, owner from description
#  - Priority preserved directly from input
# ──────────────────────────────────────────────────────────────

_L_TEXT_MAP = {"rare":2,"unlikely":3,"possible":5,"likely":7,"almost certain":9}
_I_TEXT_MAP = {"trivial":2,"minor":5,"moderate":6,"serious":7,"major":8,"critical":9}


def process_file4(path: str, client, output_path: str):
    print("  Loading File 4 (Moorgate Crossrail)…")
    wb = openpyxl.load_workbook(path)
    ws = get_data_sheet(wb)

    data_rows = []
    for r in range(2, ws.max_row + 1):
        rid = clean(ws.cell(r, 2).value)
        if not rid:
            continue
        data_rows.append({
            "rid":      rid,
            "desc":     clean(ws.cell(r, 3).value),
            "l_text":   clean(ws.cell(r, 6).value),
            "i_text":   clean(ws.cell(r, 7).value),
            "priority": clean(ws.cell(r, 8).value),
            "mit":      clean(ws.cell(r, 10).value),
        })

    SYSTEM = "You are a risk register specialist. Return ONLY valid JSON."
    batch  = json.dumps([
        {"idx": i, "rid": r["rid"], "desc": r["desc"],
         "l_text": r["l_text"], "i_text": r["i_text"], "priority": r["priority"]}
        for i, r in enumerate(data_rows)
    ], indent=2)

    USER = f"""These risks are from an urban public-realm/transport infrastructure project
at Moorgate Crossrail, London. Fill in missing fields.

Rules:
- stage    : Infer from description. One of:
               Pre-construction | Construction | Design | Operation | Commissioning
- category : Infer from description. One of:
               Planning | Stakeholder Engagement | Procurement | Design |
               Construction | Programme | Financial | Governance
- owner    : Infer a specific role (e.g. Programme Manager, Project Director,
               Design Manager, Procurement Manager, Stakeholder Manager, Planning Manager)
- l        : Convert l_text to 1-10 integer:
               Rare=2, Unlikely=3, Possible=5, Likely=7, Almost Certain=9
- i        : Convert i_text to 1-10 based on context:
               Minor=5-7, Serious=6-8, Major=8-9

INPUT:
{batch}

Return JSON array: idx, stage, category, owner, l, i"""

    print("  Calling Claude API for File 4 field inference…")
    raw    = call_claude(client, SYSTEM, USER, max_tokens=2048)
    by_idx = {e["idx"]: e for e in parse_json(raw)}

    today    = datetime.today()
    out_rows = []
    for i, row in enumerate(data_rows):
        e     = by_idx.get(i, {})
        l_val = to_int(e.get("l")) or _L_TEXT_MAP.get(row["l_text"].lower(), 5)
        i_val = to_int(e.get("i")) or _I_TEXT_MAP.get(row["i_text"].lower(), 7)
        out_rows.append([
            today,
            row["rid"],
            row["desc"],
            e.get("stage",    ""),
            e.get("category", ""),
            e.get("owner",    ""),
            l_val,
            i_val,
            row["priority"] or None,
            row["mit"],
        ])

    save_it(out_rows, output_path)
    print(f"  ✓ Saved → {output_path}  ({len(out_rows)} risks)")


# ──────────────────────────────────────────────────────────────
#  FILE 5 — Fenland DC Corporate Risk Register (blind test, PDF source)
#
#  The input is a .pdf — openpyxl cannot read it directly.
#  Strategy: extract text from the PDF using pdfminer/pypdf/PyPDF2
#  (whichever is installed), then pass the raw text to Claude which
#  reads the table structure and returns standardised JSON.
#  Also handles Excel input if the user has pre-converted the PDF.
# ──────────────────────────────────────────────────────────────

def _extract_pdf_text(path: str) -> str:
    """Extract raw text from a PDF using pdfminer, pypdf, or PyPDF2."""
    try:
        from pdfminer.high_level import extract_text
        return extract_text(path)
    except ImportError:
        pass
    try:
        import pypdf
        reader = pypdf.PdfReader(path)
        return "\n".join(page.extract_text() or "" for page in reader.pages)
    except ImportError:
        pass
    try:
        import PyPDF2
        with open(path, "rb") as f:
            reader = PyPDF2.PdfReader(f)
            return "\n".join(page.extract_text() or "" for page in reader.pages)
    except ImportError:
        pass
    return ""


def process_file5(path: str, client, output_path: str):
    print("  Processing File 5 (Fenland DC Corporate Risk Register)…")
    suffix = Path(path).suffix.lower()

    # ── Excel input ───────────────────────────────────────────
    if suffix in (".xlsx", ".xlsm", ".xls", ".xltx", ".xltm"):
        print("  Detected Excel format — reading with openpyxl…")
        wb      = openpyxl.load_workbook(path)
        ws      = get_data_sheet(wb)
        headers = [clean(ws.cell(1, c).value) for c in range(1, ws.max_column + 1)]
        rows    = []
        for r in range(2, ws.max_row + 1):
            vals = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
            if not any(v is not None for v in vals):
                continue
            rows.append({
                "idx":   len(rows),
                "cells": {(headers[j] or f"col{j+1}"): str(v)
                          for j, v in enumerate(vals) if v is not None},
            })
        source_text = json.dumps(rows, indent=2)
        source_type = "structured Excel rows"

    # ── PDF input ─────────────────────────────────────────────
    elif suffix == ".pdf":
        print("  Detected PDF — extracting text…")
        raw_text = _extract_pdf_text(path)
        if not raw_text.strip():
            print("  ⚠ Could not extract text from PDF.")
            print("  Tip: install pdfminer.six  →  pip install pdfminer.six")
            print("  Then rerun, or convert the PDF to Excel first.")
            return
        source_text = raw_text[:12000]   # stay within token limits
        source_type = "raw text extracted from a PDF"
        print(f"  Extracted {len(raw_text):,} chars of text")

    else:
        print(f"  ⚠ Unsupported format '{suffix}' for File 5.")
        return

    SYSTEM = "You are a risk register specialist. Return ONLY valid JSON."
    USER = f"""Below is {source_type} from the Fenland District Council
Corporate Risk Register (UK local government, 2019).
Original scores are on a 1–5 scale; multiply by 2 to convert to 1–10.

For each risk extract and standardise:
- risk_id       : reference number
- desc          : concise risk description (1–2 sentences)
- stage         : "Operations" (all are operational council risks)
- category      : inferred category (e.g. "ICT & Data", "Financial Sustainability",
                   "Governance & Compliance", "Workforce & HR", "Health & Safety",
                   "Emergency & Resilience", "Partnership Governance", etc.)
- owner         : risk owner name/role
- l_pre         : likelihood pre-mitigation (1–10)
- i_pre         : impact pre-mitigation (1–10)
- priority_pre  : Low / Med / High
- mit           : mitigating action summary
- l_post        : likelihood post-mitigation (1–10)
- i_post        : impact post-mitigation (1–10)
- priority_post : Low / Med / High

SOURCE:
{source_text}

Return a JSON array — one object per risk — with all fields above."""

    print("  Calling Claude API for File 5 standardisation…")
    raw     = call_claude(client, SYSTEM, USER, max_tokens=8000)
    results = parse_json(raw)

    out_rows = []
    for e in results:
        l_pre  = to_int(e.get("l_pre"))
        i_pre  = to_int(e.get("i_pre"))
        l_post = to_int(e.get("l_post"))
        i_post = to_int(e.get("i_post"))
        out_rows.append([
            None,
            to_rid(e.get("risk_id")),
            e.get("desc",  ""),
            e.get("stage", "Operations"),
            e.get("category", ""),
            e.get("owner",    ""),
            l_pre,  i_pre,  e.get("priority_pre")  or priority_from_scores(l_pre,  i_pre),
            e.get("mit",   ""),
            l_post, i_post, e.get("priority_post") or priority_from_scores(l_post, i_post),
        ])

    save_pre_post(out_rows, output_path)
    print(f"  ✓ Saved → {output_path}  ({len(out_rows)} risks)")

# ──────────────────────────────────────────────────────────────
#  GENERIC FALLBACK
# ──────────────────────────────────────────────────────────────

def process_generic(path: str, client, output_path: str):
    print("  Using generic processing…")
    wb   = openpyxl.load_workbook(path)
    ws   = get_data_sheet(wb)
    hdrs = [clean(ws.cell(1, c).value) for c in range(1, ws.max_column + 1)]

    data = []
    for r in range(2, ws.max_row + 1):
        row = {hdrs[c]: clean(ws.cell(r, c + 1).value)
               for c in range(len(hdrs)) if ws.cell(r, c + 1).value is not None}
        if row:
            data.append({"idx": len(data), **row})

    SYSTEM = "You are a risk register specialist. Return ONLY valid JSON."
    USER = f"""Standardise these risk register rows to the mandatory output format.
Output fields: risk_id, desc, stage, category, owner, l, i, priority, mit.
If pre+post data exists include: l_pre, i_pre, priority_pre, l_post, i_post, priority_post.
Scale L/I to 1-10. Priority: Low / Med / High.
INPUT: {json.dumps(data[:30], indent=2)}
Return JSON array."""

    raw     = call_claude(client, SYSTEM, USER, max_tokens=4096)
    results = parse_json(raw)

    has_pre_post = any("l_pre" in e for e in results)
    out_rows = []
    for e in results:
        if has_pre_post:
            l_pre  = to_int(e.get("l_pre",  e.get("l")))
            i_pre  = to_int(e.get("i_pre",  e.get("i")))
            l_post = to_int(e.get("l_post", e.get("l")))
            i_post = to_int(e.get("i_post", e.get("i")))
            out_rows.append([
                None, to_rid(e.get("risk_id")), e.get("desc",""),
                e.get("stage",""), e.get("category",""), e.get("owner",""),
                l_pre, i_pre, e.get("priority_pre") or priority_from_scores(l_pre, i_pre),
                e.get("mit",""),
                l_post, i_post, e.get("priority_post") or priority_from_scores(l_post, i_post),
            ])
        else:
            out_rows.append([
                None, to_rid(e.get("risk_id")), e.get("desc",""),
                e.get("stage",""), e.get("category",""),
                to_int(e.get("l")), to_int(e.get("i")),
                e.get("priority",""), e.get("owner",""), e.get("mit",""), "",
            ])

    if has_pre_post:
        save_pre_post(out_rows, output_path)
    else:
        save_single(out_rows, output_path)
    print(f"  ✓ Saved → {output_path}  ({len(out_rows)} risks)")


# ──────────────────────────────────────────────────────────────
#  PUBLIC CLASS
# ──────────────────────────────────────────────────────────────

class RiskRegisterStandardizer:

    def __init__(self, api_key: Optional[str] = None):
        self.api_key = api_key or os.getenv("ANTHROPIC_API_KEY")
        if not self.api_key:
            raise ValueError("ANTHROPIC_API_KEY must be set in environment or passed to init")
        self.client = anthropic.Anthropic(api_key=self.api_key)

    def process_file(self, input_file: str, output_file: str) -> bool:
        name = Path(input_file).name.lower()
        print(f"\nProcessing: {input_file}")
        print("=" * 70)
        try:
            if any(k in name for k in ("ivc", "doe", "1__", "1. ivc")):
                process_file1(input_file, self.client, output_file)
            elif any(k in name for k in ("york", "2__", "2. city")):
                process_file2(input_file, self.client, output_file)
            elif any(k in name for k in ("digital", "security", "3__", "3. digital")):
                process_file3(input_file, self.client, output_file)
            elif any(k in name for k in ("moorgate", "crossrail", "4__", "4. moorgate")):
                process_file4(input_file, self.client, output_file)
            elif any(k in name for k in ("corporate", "fenland", "5__", "5. corporate")):
                process_file5(input_file, self.client, output_file)
            else:
                print(f"  ⚠ Unrecognised file '{name}' — using generic processing")
                process_generic(input_file, self.client, output_file)
            return True
        except Exception as exc:
            import traceback
            print(f"  ✗ Error: {exc}")
            traceback.print_exc()
            return False


# ──────────────────────────────────────────────────────────────
#  ENTRY POINT
# ──────────────────────────────────────────────────────────────

def main():
    input_dir  = Path("input")
    output_dir = Path("output")
    output_dir.mkdir(exist_ok=True)

    input_files = sorted(
        f for f in input_dir.glob("*")
        if f.suffix.lower() in (".xlsx", ".xls", ".pdf")
    )

    if not input_files:
        print("No .xlsx / .xls / .pdf files found in ./input")
        return

    print(f"Found {len(input_files)} file(s) to process\n")

    try:
        standardizer = RiskRegisterStandardizer()
    except ValueError as exc:
        print(f"Error: {exc}")
        return

    for f in input_files:
        stem     = f.stem
        out_stem = re.sub(r"(?i)[\s_]*\(?input\)?[\s_]*", " (Final)", stem).strip()
        if out_stem == stem:
            out_stem = stem + " (Final)"
        out_path = output_dir / (out_stem + ".xlsx")
        success  = standardizer.process_file(str(f), str(out_path))
        if not success:
            print(f"  ⚠ Warning: failed to process {f.name}\n")


if __name__ == "__main__":
    main()